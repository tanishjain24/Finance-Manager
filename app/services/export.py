"""Export helpers for CSV, Excel, and PDF reports."""
import csv
import io
from datetime import datetime
from flask import make_response


def export_transactions_csv(transactions, filename='transactions'):
	output = io.StringIO()
	writer = csv.writer(output)
	writer.writerow(['Date', 'Name', 'Category', 'Account', 'Amount', 'Notes', 'Recurring'])
	for item in transactions:
		t = item['t']
		writer.writerow([
			t.timestamp.strftime('%Y-%m-%d %H:%M'),
			t.transaction_name,
			item.get('category_name', ''),
			item.get('account_name', ''),
			float(t.amount),
			t.note or '',
			'Yes' if t.recurring else 'No',
		])
	output.seek(0)
	response = make_response(output.getvalue())
	response.headers['Content-Disposition'] = 'attachment; filename={}.csv'.format(filename)
	response.headers['Content-Type'] = 'text/csv'
	return response


def export_transactions_excel(transactions, filename='transactions'):
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font, PatternFill, Alignment
	except ImportError:
		return export_transactions_csv(transactions, filename)

	wb = Workbook()
	ws = wb.active
	ws.title = 'Transactions'
	headers = ['Date', 'Name', 'Category', 'Account', 'Amount', 'Notes', 'Recurring']
	header_font = Font(bold=True, color='FFFFFF')
	header_fill = PatternFill(start_color='10B981', end_color='10B981', fill_type='solid')

	for col, header in enumerate(headers, 1):
		cell = ws.cell(row=1, column=col, value=header)
		cell.font = header_font
		cell.fill = header_fill
		cell.alignment = Alignment(horizontal='center')

	for row_idx, item in enumerate(transactions, 2):
		t = item['t']
		ws.cell(row=row_idx, column=1, value=t.timestamp.strftime('%Y-%m-%d %H:%M'))
		ws.cell(row=row_idx, column=2, value=t.transaction_name)
		ws.cell(row=row_idx, column=3, value=item.get('category_name', ''))
		ws.cell(row=row_idx, column=4, value=item.get('account_name', ''))
		ws.cell(row=row_idx, column=5, value=float(t.amount))
		ws.cell(row=row_idx, column=6, value=t.note or '')
		ws.cell(row=row_idx, column=7, value='Yes' if t.recurring else 'No')

	for col in ws.columns:
		max_len = max(len(str(cell.value or '')) for cell in col)
		ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)
	response = make_response(buf.getvalue())
	response.headers['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(filename)
	response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
	return response


def export_transactions_pdf(transactions, user, filename='report'):
	try:
		from reportlab.lib import colors
		from reportlab.lib.pagesizes import letter
		from reportlab.lib.styles import getSampleStyleSheet
		from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
	except ImportError:
		return export_transactions_csv(transactions, filename)

	buf = io.BytesIO()
	doc = SimpleDocTemplate(buf, pagesize=letter)
	elements = []
	styles = getSampleStyleSheet()

	elements.append(Paragraph('Finance Manager — Transaction Report', styles['Title']))
	elements.append(Paragraph('Generated: {} | User: {} {}'.format(
		datetime.utcnow().strftime('%Y-%m-%d'), user.first_name, user.last_name),
		styles['Normal']))
	elements.append(Spacer(1, 20))

	data = [['Date', 'Name', 'Category', 'Amount']]
	for item in transactions[:100]:
		t = item['t']
		data.append([
			t.timestamp.strftime('%Y-%m-%d'),
			t.transaction_name[:20],
			item.get('category_name', '')[:15],
			'{:.2f}'.format(float(t.amount)),
		])

	table = Table(data, repeatRows=1)
	table.setStyle(TableStyle([
		('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10B981')),
		('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
		('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
		('FONTSIZE', (0, 0), (-1, -1), 9),
		('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8FAFC')]),
		('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
		('ALIGN', (3, 1), (3, -1), 'RIGHT'),
	]))
	elements.append(table)
	doc.build(elements)
	buf.seek(0)

	response = make_response(buf.getvalue())
	response.headers['Content-Disposition'] = 'attachment; filename={}.pdf'.format(filename)
	response.headers['Content-Type'] = 'application/pdf'
	return response
